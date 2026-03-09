# exporter.py - Generate formatted Excel output matching the ABB report style

import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import date


# ─── COLOURS ──────────────────────────────────────────────────────────────────
ORANGE_FILL   = PatternFill("solid", fgColor="C8501E")
ORANGE_LT     = PatternFill("solid", fgColor="FAD7A0")
GREEN_FILL    = PatternFill("solid", fgColor="92D050")
RED_FILL      = PatternFill("solid", fgColor="FF6B6B")
GREY_FILL     = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
BEIGE_FILL    = PatternFill("solid", fgColor="FFF2CC")

BOLD_WHITE    = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
BOLD_BLACK    = Font(bold=True, color="000000", name="Calibri", size=10)
NORMAL        = Font(name="Calibri", size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center")
RIGHT         = Alignment(horizontal="right",  vertical="center")

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT_FMT    = "0.000%"
NUMBER_FMT = "#,##0.00"
INT_FMT    = "0"


def generate_excel(
    report_date: date,
    periodicity: str,
    category: str,
    t1: pd.DataFrame,
    t2: pd.DataFrame,
    t3: pd.DataFrame,
) -> bytes:
    """
    Generate a formatted Excel workbook with 3 sheets (Fonds, Synthèse, Ranking).
    Returns bytes for st.download_button.
    """
    wb = Workbook()
    _build_sheet1(wb, t1, category, report_date, periodicity)
    _build_sheet2(wb, t2, category)
    _build_sheet3(wb, t3, category)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── SHEET 1: Fonds ───────────────────────────────────────────────────────────

def _build_sheet1(wb: Workbook, t1: pd.DataFrame, category: str,
                  report_date: date, periodicity: str):
    ws = wb.create_sheet("Tableau 1 - Fonds")

    col_headers = [
        "Code ISIN", "OPCVM", "Société de Gestion", "Souscripteurs",
        "Périodicité VL", "Actif Net", "Classification",
        "Performance quotidienne", "YTD",
    ]
    col_keys = [
        "isin", "opcvm", "societe_gestion", "souscripteurs",
        "periodicite", "actif_net", "classification", "perf_1j", "ytd",
    ]
    col_widths = [18, 35, 28, 14, 14, 22, 14, 22, 12]

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"ABB – OPCVM {category} – {periodicity.upper()} – {report_date.strftime('%d/%m/%Y')}"
    )
    title_cell.fill = ORANGE_FILL
    title_cell.font = BOLD_WHITE
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 22

    # Header row
    for c_idx, header in enumerate(col_headers, 1):
        cell = ws.cell(row=2, column=c_idx, value=header)
        cell.fill = ORANGE_LT
        cell.font = BOLD_BLACK
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[2].height = 30

    # Data rows
    for r_idx, (_, row) in enumerate(t1.iterrows(), 3):
        fill = WHITE_FILL if r_idx % 2 == 0 else GREY_FILL
        for c_idx, key in enumerate(col_keys, 1):
            val = row.get(key, "")
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = BORDER
            cell.alignment = CENTER

            if key in ("perf_1j", "ytd"):
                num = _safe_float(val)
                if num is not None:
                    cell.value = num / 100  # store as decimal for Excel %
                    cell.number_format = "0.000%"
                    cell.fill = GREEN_FILL if num > 0 else (RED_FILL if num < 0 else fill)
                else:
                    cell.value = ""
                    cell.fill = fill
            elif key == "actif_net":
                num = _safe_float(val)
                cell.value = num if num is not None else ""
                cell.number_format = "#,##0.00"
                cell.fill = fill
            else:
                cell.value = str(val) if val else ""
                cell.fill = fill
                cell.alignment = LEFT if key in ("opcvm", "societe_gestion") else CENTER
            cell.font = NORMAL

    # Column widths
    for c_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws.freeze_panes = "A3"


# ─── SHEET 2: Synthèse ────────────────────────────────────────────────────────

def _build_sheet2(wb: Workbook, t2: pd.DataFrame, category: str):
    ws = wb.create_sheet("Tableau 2 - Synthese")

    # Title
    ws.merge_cells("A1:C1")
    cell = ws["A1"]
    cell.value = f"{category} MARCHÉ"
    cell.fill = ORANGE_FILL
    cell.font = BOLD_WHITE
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 22

    for r_idx, (_, row) in enumerate(t2.iterrows(), 2):
        indicateur = row.get("Indicateur", "")
        valeur     = row.get("Valeur", "")
        detail     = row.get("Détail", "")

        fill = ORANGE_LT if r_idx == 2 else (GREY_FILL if r_idx % 2 == 0 else WHITE_FILL)

        cell_a = ws.cell(row=r_idx, column=1, value=indicateur)
        cell_a.font  = BOLD_BLACK if r_idx <= 2 else NORMAL
        cell_a.fill  = fill
        cell_a.alignment = LEFT
        cell_a.border = BORDER

        cell_b = ws.cell(row=r_idx, column=2, value=valeur)
        cell_b.alignment = CENTER
        cell_b.border = BORDER
        # Colour by sign
        if isinstance(valeur, str) and "%" in valeur:
            sign = _sign_of_pct_str(valeur)
            cell_b.fill = GREEN_FILL if sign > 0 else (RED_FILL if sign < 0 else fill)
        else:
            cell_b.fill = fill
        cell_b.font = BOLD_BLACK if r_idx <= 3 else NORMAL

        cell_c = ws.cell(row=r_idx, column=3, value=detail)
        cell_c.font = Font(bold=True, color="006400", name="Calibri", size=10)
        cell_c.fill = fill
        cell_c.alignment = LEFT
        cell_c.border = BORDER

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30


# ─── SHEET 3: Ranking ─────────────────────────────────────────────────────────

def _build_sheet3(wb: Workbook, t3: pd.DataFrame, category: str):
    ws = wb.create_sheet("Tableau 3 - Ranking")

    if t3.empty:
        ws["A1"] = "Aucune donnée de classement disponible."
        return

    headers = list(t3.columns)
    col_widths = {
        "OPCVM": 30,
        "Rang interne": 13,
        "Rang marché": 13,
        "Position": 12,
        "Écart vs Plus perf.": 18,
        "Écart vs Moyenne": 16,
        "Écart vs Moins perf.": 20,
        "YTD": 12,
        "YTD Marché Moyen": 16,
        "Écart vs Marché": 16,
        "Note /10": 10,
    }

    pct_cols = {
        "Écart vs Plus perf.", "Écart vs Moyenne", "Écart vs Moins perf.",
        "YTD", "YTD Marché Moyen", "Écart vs Marché",
    }

    # Header
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = ORANGE_FILL
        cell.font = BOLD_WHITE
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    # Data
    for r_idx, (_, row) in enumerate(t3.iterrows(), 2):
        base_fill = WHITE_FILL if r_idx % 2 == 0 else GREY_FILL

        for c_idx, col in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val  = row[col]
            cell.border = BORDER

            if col in pct_cols:
                num = _safe_float(val)
                if num is not None:
                    cell.value = num / 100
                    cell.number_format = "0.000%"
                    # Colour logic
                    if col in ("Écart vs Plus perf.", "Écart vs Moins perf."):
                        # Always >= 0 — neutral orange tint
                        cell.fill = BEIGE_FILL
                    elif col in ("Écart vs Moyenne", "Écart vs Marché"):
                        cell.fill = GREEN_FILL if num > 0 else (RED_FILL if num < 0 else base_fill)
                    elif col in ("YTD", "YTD Marché Moyen"):
                        cell.fill = GREEN_FILL if num > 0 else (RED_FILL if num < 0 else base_fill)
                    else:
                        cell.fill = base_fill
                else:
                    cell.value = ""
                    cell.fill = base_fill
            elif col == "Note /10":
                cell.value = int(val) if pd.notna(val) else ""
                note = int(val) if pd.notna(val) else 0
                cell.fill = GREEN_FILL if note == 10 else (BEIGE_FILL if note == 5 else RED_FILL)
                cell.alignment = CENTER
            elif col == "Position":
                cell.value = str(val)
                if val == "Top 25%":
                    cell.fill = GREEN_FILL
                elif val == "Bas 25%":
                    cell.fill = RED_FILL
                else:
                    cell.fill = BEIGE_FILL
                cell.alignment = CENTER
            else:
                cell.value = str(val) if pd.notna(val) else ""
                cell.fill = base_fill
                cell.alignment = LEFT if col == "OPCVM" else CENTER

            cell.font = BOLD_BLACK if col == "OPCVM" else NORMAL

    # Column widths
    for c_idx, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col, 14)

    ws.freeze_panes = "A2"


# ─── UTILS ────────────────────────────────────────────────────────────────────

def _safe_float(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return None if np.isnan(val) else float(val)
    try:
        return float(str(val).replace(",", ".").replace("%", "").strip())
    except Exception:
        return None


def _sign_of_pct_str(s: str) -> int:
    """Return 1, -1, or 0 from a formatted pct string like '+1.23%'."""
    s = s.replace("%", "").replace(",", ".").strip()
    try:
        v = float(s)
        return 1 if v > 0 else (-1 if v < 0 else 0)
    except Exception:
        return 0
